import gradio as gr
import requests
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
from rouge_score import rouge_scorer
from transformers import MarianMTModel, MarianTokenizer
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Evaluate translation quality using BLEU score
def evaluate_bleu_score(reference, translation):
    reference = [reference.split()]  # Split the reference translation into words
    translation = translation.split()  # Split the translation into words
    
    smoothie = SmoothingFunction().method4
    bleu_score = sentence_bleu(reference, translation, smoothing_function=smoothie)
    
    return bleu_score

# Evaluate translation quality using ROUGE score
def evaluate_rouge_score(reference, translation):
    scorer = rouge_scorer.RougeScorer(['rouge1', 'rouge2', 'rougeL'], use_stemmer=True)
    scores = scorer.score(reference, translation)
    return scores

# Fetch reference translation from MyMemory API
def get_reference_translation(text, source_lang='tr', target_lang='en'):
    url = "https://api.mymemory.translated.net/get"
    params = {
        'q': text,
        'langpair': f'{source_lang}|{target_lang}'
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        data = response.json()
        if 'responseData' in data and 'translatedText' in data['responseData']:
            return data['responseData']['translatedText']
        else:
            print("Failed to get a valid translation from MyMemory API. You might want to consider an alternative source.")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error occurred during API request: {e}")
    except ValueError:
        print("Error occurred while parsing the JSON response.")
    return None

# Translate the input text using two models and calculate evaluation scores
def translate_and_evaluate(text):
    model_name_1 = 'Helsinki-NLP/opus-mt-tc-big-tr-en'
    model_1 = MarianMTModel.from_pretrained(model_name_1)
    tokenizer_1 = MarianTokenizer.from_pretrained(model_name_1)

    model_name_2 = 'Helsinki-NLP/opus-mt-tr-en'
    model_2 = MarianMTModel.from_pretrained(model_name_2)
    tokenizer_2 = MarianTokenizer.from_pretrained(model_name_2)

    cost_per_token_model_1 = 0.0004
    cost_per_token_model_2 = 0.0003

    models = [model_1, model_2]
    tokenizers = [tokenizer_1, tokenizer_2]
    costs = [cost_per_token_model_1, cost_per_token_model_2]

    reference_translation = get_reference_translation(text)
    if not reference_translation:
        return None, None

    all_results = []
    for idx, (model, tokenizer, cost_per_token) in enumerate(zip(models, tokenizers, costs)):
        start_time = time.time()
        tokens = tokenizer(text, return_tensors="pt", padding=True)
        num_tokens = len(tokens.input_ids[0])
        translated = model.generate(**tokens)
        translated_text = tokenizer.batch_decode(translated, skip_special_tokens=True)[0]
        end_time = time.time()
        time_taken = end_time - start_time
        cost = cost_per_token * num_tokens

        bleu_score = evaluate_bleu_score(reference_translation, translated_text)
        rouge_score = evaluate_rouge_score(reference_translation, translated_text)

        all_results.append({
            'Input': text,
            'Model': f'Model {idx + 1}',
            'Translation': translated_text,
            'Reference Translation': reference_translation,  # Added reference translation
            'BLEU': bleu_score,
            'ROUGE-1': rouge_score['rouge1'].fmeasure,
            'ROUGE-2': rouge_score['rouge2'].fmeasure,
            'ROUGE-L': rouge_score['rougeL'].fmeasure,
            'Time': time_taken,
            'Cost': cost
        })

    # Save results to an Excel file
    df = pd.DataFrame(all_results)
    excel_path = 'translation_results.xlsx'
    df.to_excel(excel_path, index=False)

    # Highlight the row with the best model
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find the best model
    best_result = max(all_results, key=lambda x: (x['BLEU'], x['ROUGE-1']))
    best_model_idx = next(i for i, result in enumerate(all_results) if result['Model'] == f'Model {all_results.index(best_result) + 1}')

    # Apply red fill to the best model row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(df.columns) + 1):
        ws.cell(row=best_model_idx + 2, column=col).fill = red_fill  # +2 because Excel is 1-indexed and we have a header row

    wb.save(excel_path)

    best_translation = best_result['Translation']
    cost_info = f"${best_result['Cost']:.4f}"
    return best_translation, cost_info

# Gradio interface
with gr.Blocks() as demo:
    chatbox = gr.Chatbot(label="Translation Chatbot", height=500)
    input_text = gr.Textbox(label="Enter Turkish Text")
    translate_button = gr.Button("Translate")
    
    def chat(user_input, history):
        translation, cost = translate_and_evaluate(user_input)
        history.append((user_input, f"{translation} (Cost: {cost})"))
        return history, ""
    
    translate_button.click(chat, [input_text, chatbox], [chatbox, input_text])

demo.launch()
